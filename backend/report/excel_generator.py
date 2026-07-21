import openpyxl


def _write_count_sheet(wb, title: str, label_column: str, counts: dict) -> None:
    sheet = wb.create_sheet(title)
    sheet.append([label_column, "Số lượng"])
    for key, count in counts.items():
        sheet.append([str(key), count])


def generate_excel(date_from, date_to, aggregates: dict, output_path: str) -> None:
    # 1 sheet riêng cho từng bảng thống kê (giữ đúng cấu trúc như docx_generator.py) +
    # 1 sheet "Bài viết" liệt kê chi tiết — dùng openpyxl thuần, không cần template file.
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # bỏ sheet mặc định trống

    info_sheet = wb.create_sheet("Tổng quan")
    info_sheet.append(["Báo cáo NGS Monitor"])
    info_sheet.append(["Khoảng thời gian", f"{date_from} – {date_to}"])

    _write_count_sheet(wb, "Theo cơ quan", "Cơ quan", aggregates["source_counts"])
    _write_count_sheet(wb, "Theo chủ đề", "Chủ đề", aggregates["topic_counts"])
    _write_count_sheet(wb, "Top từ khóa", "Từ khóa", aggregates["keyword_counts"])
    _write_count_sheet(wb, "Theo tháng", "Tháng", aggregates["monthly_counts"])
    _write_count_sheet(wb, "Sentiment", "Sentiment", aggregates["sentiment_counts"])
    _write_count_sheet(wb, "Emotion", "Emotion", aggregates["emotion_counts"])
    _write_count_sheet(wb, "Tổng hợp", "Chỉ số", aggregates["summary_stats"])

    articles_sheet = wb.create_sheet("Bài viết")
    articles_sheet.append(["Tiêu đề", "URL", "Sentiment", "Emotion", "Confidence", "Needs review"])
    for article in aggregates["articles"]:
        articles_sheet.append(
            [
                article["title"] or "",
                article["url"] or "",
                article["sentiment"] or "",
                article["emotion"] or "",
                article["confidence"],
                article["needs_review"],
            ]
        )

    wb.save(output_path)
