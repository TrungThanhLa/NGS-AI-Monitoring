import os
import uuid

import openpyxl

from backend.report.excel_generator import generate_excel


def test_generate_excel_creates_readable_workbook_with_expected_sheets(tmp_path):
    aggregates = {
        "articles": [{"title": "Bài 1", "url": "https://vtv.vn/a1", "sentiment": "negative",
                      "emotion": "Fear", "confidence": 0.9, "needs_review": False}],
        "sentiment_counts": {"negative": 1},
        "emotion_counts": {"Fear": 1},
        "source_counts": {"VTV": 1},
        "topic_counts": {"A": 1},
        "keyword_counts": {"kw1": 1},
        "monthly_counts": {"2026-06": 1},
        "summary_stats": {"Tổng số bài": 1},
    }
    output_path = str(tmp_path / f"{uuid.uuid4()}.xlsx")

    generate_excel("2026-06-01", "2026-06-30", aggregates, output_path)

    assert os.path.exists(output_path)
    wb = openpyxl.load_workbook(output_path)
    assert "Bài viết" in wb.sheetnames
    assert "Tổng hợp" in wb.sheetnames
    articles_sheet = wb["Bài viết"]
    assert articles_sheet.cell(row=1, column=1).value == "Tiêu đề"
    assert articles_sheet.cell(row=2, column=1).value == "Bài 1"
